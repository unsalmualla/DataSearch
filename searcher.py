# Excell dosyasından veri okuyabilmek için openpyxl
# modülünden load_workbook fonksiyonunu içeriye aktarıyoruz.
from openpyxl import load_workbook


# Programımız main fonksiyonu içerisinde çaışacak.
def main(excellFile, newExcellFile, sheetOneV, sheetTwoV):
    print("Loading data from " + excellFile)
    book = load_workbook(excellFile)
    print("Upload complete")
    sheetOne = book.worksheets[int(sheetOneV)]
    sheetTwo = book.worksheets[int(sheetTwoV)]

    # Gen değeri alınacak kişilere tek tek ulaşmak için
    # bir döngü oluşturulur
    for colunm in range(2, sheetTwo.max_column + 1):
        # Kişi isimlerine ulaşabilmek için, o sutünda ki değer alınır.
        personOne = sheetTwo.cell(row=1, column=colunm).value
        # Eğer değer boş yani None ise, program sonlanır.
        if personOne != None:
            print("Person information is placed: " + str(personOne))

            # Kişinin gen değerlerine tek tek diğer sayfada aramak için ulaşılır.
            for row in range(2, sheetTwo.max_row + 1):
                # Kişinin gen ismi alınır.
                geneOne = sheetTwo.cell(row=row, column=1).value
                print("Searched Gene: " + str(geneOne))

                # Belirlenen kişi ismi, birinci sayfada ki ilk satırda sütun sütun aranır.
                for i in range(2, sheetOne.max_column + 1):
                    # Her sutünda ki değer <isim1> personTwo değerine aktarılır.
                    personTwo = sheetOne.cell(row=1, column=i).value

                    # Eğer ilgili sütunda ki kişi ile aranan kişi uyuyorsa
                    # Gen aranmaya başlanır.
                    if personTwo == personOne:
                        # Bulunan kişi için ilgili gen aranır.
                        for j in range(2, sheetOne.max_row + 1):

                            geneTwo = sheetOne.cell(row=j, column=1).value
                            # Eğer gen değeri de tutarsa...
                            if geneTwo == geneOne:
                                # Satır ve Sütun kesişimi noktada bizim gen değerimiz bulunur. Değer alınır.
                                walle = sheetOne.cell(row=j, column=i).value
                                print("Found: " + str(walle) + "| " + str(j) + "/" + str(i))
                                # Bu ikinci sayfada, genini aradığımız kişinin gen değerine yazılır.
                                sheetTwo.cell(row=row, column=colunm).value = walle
                            else:
                                pass
                    else:
                        pass
        else:
            break


    # Yeni ismi ile excell dosyasını kaydediyoruz.
    book.save(newExcellFile)
#Program çalıştırıldığında kullanıcıdan istenen dosyalar ve sayfa numaraları.
excellFile = input("İşlemin yapılacağı dosyayı giriniz: ")
newExcellFile = input("Yeni Oluşturulacak Excell Dosyanın adı: ")
sheetOneV = input("mRNA zscores'un bulunduğu sheet sıra numarasını giriniz: ")
sheetTwoV = input("Dna metilasyon verisinin bulunduğu sheetin sıra numarasını giriniz: ")
print("-" * 50)
print("program başlıyor")

# Aldığımız verileri main fonksiyonuna gönderiyoruz.
main(excellFile, newExcellFile, sheetOneV, sheetTwoV)
